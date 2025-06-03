import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment
import os
import platform
import subprocess

class FilterWithTimeApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Data Analysis - ©TT Nguyen 2025. All rights reserved.")
        self.root.geometry("600x450")
        self.df = None
        self.filename = None
        self.last_save_path = None

        # Dùng font Segoe UI, bình thường, size 11 cho tất cả widget
        self.style = ttk.Style()
        self.style.configure('TLabel', font=("Segoe UI", 11, "normal"))
        self.style.configure('TEntry', font=("Segoe UI", 11, "normal"))
        self.style.configure('TCombobox', font=("Segoe UI", 11, "normal"))
        self.style.configure('TButton', font=("Segoe UI", 11, "normal"))

        # === Step 1: Select Data File ===
        file_frame = ttk.Labelframe(root, text="Step 1: Select Data File", padding=(15, 10))
        file_frame.pack(padx=15, pady=(15, 8), fill="x")
        ttk.Button(file_frame, text="📁 Choose File", command=self.load_file,
                   bootstyle="outline-danger", width=20).pack()

        # === Step 2: Filtering Options ===
        filter_frame = ttk.Labelframe(root, text="Step 2: Filtering Options", padding=(15, 10))
        filter_frame.pack(padx=15, pady=8, fill="x")

        ttk.Label(filter_frame, text="Select Filter Column:").grid(row=0, column=0, sticky="e", padx=5, pady=5)
        self.filter_col_var = tk.StringVar()
        self.filter_col_dropdown = ttk.Combobox(filter_frame, textvariable=self.filter_col_var, state="readonly", width=28)
        self.filter_col_dropdown.grid(row=0, column=1, columnspan=3, sticky="w", padx=5, pady=5)

        ttk.Label(filter_frame, text="From:").grid(row=1, column=0, sticky="e", padx=5, pady=5)
        self.from_entry = ttk.Entry(filter_frame, width=12)
        self.from_entry.grid(row=1, column=1, sticky="w", padx=(5, 20), pady=5)

        ttk.Label(filter_frame, text="To:").grid(row=1, column=2, sticky="e", padx=5, pady=5)
        self.to_entry = ttk.Entry(filter_frame, width=12)
        self.to_entry.grid(row=1, column=3, sticky="w", padx=5, pady=5)

        # === Step 3: Export Result ===
        action_frame = ttk.Labelframe(root, text="Step 3: Export Result", padding=(15, 10))
        action_frame.pack(padx=15, pady=8, fill="x")

        ttk.Button(action_frame, text="📤 Filter and Export to Excel", command=self.filter_and_export,
                   bootstyle="outline-secondary", width=30).pack(pady=(0, 10))

        # Buttons open file / folder
        button_frame = ttk.Frame(action_frame)
        button_frame.pack(fill="x")

        ttk.Button(button_frame, text="📂 Open File", command=self.open_last_file,
                   bootstyle="outline-success", width=20).pack(side="left", expand=True, fill="x", padx=(0, 7))
        ttk.Button(button_frame, text="📁 Open Folder", command=self.open_containing_folder,
                   bootstyle="outline-info", width=20).pack(side="right", expand=True, fill="x", padx=(7, 0))

        # Progress percentage label above progress bar
        self.progress_label = ttk.Label(root, text="0%", font=("Segoe UI", 10, "bold"))
        self.progress_label.pack(padx=15, pady=(5, 0))  # Label phía trên

        # Progress bar
        self.style.configure("custom.Horizontal.TProgressbar", thickness=15)
        self.progress = ttk.Progressbar(root, orient="horizontal", length=560, mode="determinate",
                                        bootstyle="striped light", style="custom.Horizontal.TProgressbar")
        self.progress.pack(pady=(0, 10), padx=15)

        # Status bar
        self.status_var = tk.StringVar(value="No data loaded.")
        status_bar = ttk.Label(root, textvariable=self.status_var, anchor="w",
                               font=("Segoe UI", 10, "normal"), bootstyle="secondary")
        status_bar.pack(fill="x", side="bottom", ipady=3)

    def load_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("All files", "*.*")])
        if not file_path:
            return

        try:
            with open(file_path, encoding="ISO-8859-1", errors="replace") as f:
                lines = f.readlines()

            data_start = next(i for i, line in enumerate(lines) if '***DATA***' in line)
            header_line = lines[data_start + 1].strip()
            data_lines = lines[data_start + 2:]

            columns = header_line.split('\t')
            data = [line.strip().split('\t') for line in data_lines if line.strip()]
            self.df = pd.DataFrame(data, columns=columns).apply(pd.to_numeric, errors='coerce')
            self.filename = file_path

            self.filter_col_dropdown['values'] = list(self.df.columns)
            default_col = "Cell Current (A)" if "Cell Current (A)" in self.df.columns else self.df.columns[0]
            self.filter_col_dropdown.set(default_col)

            self.status_var.set(f"Loaded: {file_path}")
            messagebox.showinfo("Success", "File loaded successfully!")
        except Exception as e:
            self.status_var.set("Error loading file.")
            messagebox.showerror("Error", str(e))

    def update_progress(self, percent):
        self.progress["value"] = percent
        self.progress_label.config(text=f"{int(percent)}%")
        self.root.update_idletasks()

    def filter_and_export(self):
        if self.df is None:
            messagebox.showwarning("No Data", "Please load a file first.")
            return

        try:
            low = float(self.from_entry.get())
            high = float(self.to_entry.get())
            filter_col = self.filter_col_var.get()
        except ValueError:
            messagebox.showerror("Invalid Input", "Please enter numeric values for filter range.")
            return

        time_col = "Time  (Sec)"
        hour_col = "Time (Hour)"
        common_cols = [
            "Cell Power(W)", "Anode O2 MFM Flow(sccm)",
            "Cathode H2 MFM Flow(sccm)", "Anode H2 Sensor(%)",
            "Cathode O2 Sensor(%)"
        ]
        extra_cols = [col for col in ["Cell Current (A)", "Cell Voltage(V)"] if col in self.df.columns]
        output_columns = [time_col] + extra_cols + [hour_col] + common_cols

        missing = [col for col in [time_col, filter_col] + common_cols if col not in self.df.columns]
        if missing:
            messagebox.showerror("Missing Columns", f"Required columns missing: {', '.join(missing)}")
            return

        try:
            self.update_progress(10)
            self.status_var.set("Filtering data...")

            filtered_df = self.df[(self.df[filter_col] >= low) & (self.df[filter_col] <= high)]
            filtered_df[hour_col] = (filtered_df[time_col] - filtered_df[time_col].iloc[0]) / 3600

            selected_df = filtered_df[output_columns]
            mean_row = selected_df.mean(numeric_only=True).round(3)
            mean_row[time_col] = None
            mean_row[hour_col] = None
            result_df = pd.concat([pd.DataFrame([mean_row]), selected_df], ignore_index=True)

            save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
            if save_path:
                result_df.to_excel(save_path, index=False)
                self.update_progress(60)
                self.status_var.set("Formatting Excel...")

                wb = load_workbook(save_path)
                ws = wb.active

                fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
                for cell in ws[2]:
                    cell.fill = fill
                ws.cell(row=2, column=1).comment = Comment("Average value", "System")

                header_font = Font(bold=True, name='Calibri')
                center_align = Alignment(horizontal="center", vertical="center")
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))

                for cell in ws[1]:
                    cell.font = header_font
                    cell.alignment = center_align
                    cell.border = thin_border

                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        cell.font = Font(name='Calibri')
                        cell.alignment = center_align
                        cell.border = thin_border

                for col in ws.columns:
                    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                    ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

                wb.save(save_path)

                self.last_save_path = os.path.abspath(save_path)
                self.update_progress(100)
                self.status_var.set(f"File saved: {save_path}")
                messagebox.showinfo("Done", "Excel file saved successfully!")
        except Exception as e:
            self.status_var.set("Export error.")
            messagebox.showerror("Error", str(e))
        finally:
            self.update_progress(0)

    def open_last_file(self):
        if self.last_save_path and os.path.isfile(self.last_save_path):
            try:
                if platform.system() == "Windows":
                    os.startfile(self.last_save_path)
                elif platform.system() == "Darwin":
                    subprocess.call(["open", self.last_save_path])
                else:
                    subprocess.call(["xdg-open", self.last_save_path])
            except Exception as e:
                messagebox.showerror("Error", f"Cannot open file: {e}")
        else:
            messagebox.showwarning("No File", "No recently saved file to open.")

    def open_containing_folder(self):
        if self.last_save_path and os.path.isfile(self.last_save_path):
            folder_path = os.path.dirname(os.path.abspath(self.last_save_path))
            try:
                if platform.system() == "Windows":
                    subprocess.Popen(["explorer", "/select,", os.path.normpath(self.last_save_path)])
                elif platform.system() == "Darwin":
                    subprocess.call(["open", folder_path])
                else:
                    subprocess.call(["xdg-open", folder_path])
            except Exception as e:
                messagebox.showerror("Error", f"Cannot open folder: {e}")
        else:
            messagebox.showwarning("No File", "No recently saved file to open folder.")

if __name__ == "__main__":
    root = ttk.Window(themename="journal")  # theme dịu nhẹ
    app = FilterWithTimeApp(root)
    root.mainloop()
